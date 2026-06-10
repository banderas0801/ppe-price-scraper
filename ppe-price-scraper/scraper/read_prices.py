"""
read_prices.py — Đọc kết quả scrape từ GitHub raw và điền vào file Excel gốc

Usage:
  python read_prices.py \
    --github_url https://raw.githubusercontent.com/YOUR_REPO/main/data/prices.json \
    --excel_in  "3__SỐ_LƯỢNG___PHÂN_LOẠI_CHI_TIẾT__1_.xlsx" \
    --excel_out "BaoGia_PPE_VSIP_2024_Final.xlsx"
"""

import requests
import json
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def fetch_prices(github_raw_url):
    """Fetch price data from GitHub raw URL"""
    r = requests.get(github_raw_url, timeout=15)
    r.raise_for_status()
    return r.json()


def fill_excel(price_data, excel_in, excel_out):
    """Fill prices into the original Excel file"""
    wb = openpyxl.load_workbook(excel_in)
    ws = wb["2023"]

    thin = Side(style="thin", color="BBBBBB")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    FONT = Font(name="Arial", size=9)
    GREEN_FILL = PatternFill("solid", start_color="E8F5E9")
    YELLOW_FILL = PatternFill("solid", start_color="FFF9C4")

    products = price_data.get("products", {})

    # Map product name keywords → product IDs
    NAME_MAP = {
        "mũ an toàn công nghiệp (giám sát": "mu_sseda_mat_vuong_trang",
        "mũ an toàn công nghiệp  (công nhân": "mu_sseda_mat_tron_vang",
        "mũ vải chống nắng": "mu_vai_chong_nang",
        "mũ vải an ninh thường": "mu_vai_an_ninh_thuong",
        "tấm kính chụp mặt": "tang_kính_chup_mat",
        "khẩu trang lọc bụi 3m-9914k": None,  # hợp quy - skip
        "khẩu trang chống bụi 3m -9542": None,
        "mặt nạ phòng độc 3m": None,
        "đầu lọc của mặt nạ": None,
        "găng tay vải công nhân": "gang_vai_cong_nhan",
        "gang tay vải cứu hộ": None,  # already has price
        "găng tay cao su": "gang_cao_su_dai_tay",
        "găng tay hóa chất": "gang_hoa_chat_dai_tay",
        "găng tay chống cắt": "gang_chong_cat_3m_c5",
        "găng tay y tế": "gang_y_te_cimax",
        "giày bảo hộ lao động giám sát": "giay_safety_jogger_x11",
        "giày bảo hộ lao động công nhân": None,  # hợp quy sami
        "giày vải": "giay_vai_asia",
        "ủng bảo hộ (đế, mũi thép)": None,
        "ủng nhựa": "ung_nhua_trang",
        "giày batta pccc": "giay_batta_pccc",
        "giày an ninh thường": "giay_an_ninh_thuong",
        "giày an ninh chuyên dụng": "giay_an_ninh_cd",
        "quần áo bảo hộ lao động công nhân": "qa_cong_nhan",
        "quần áo bảo hộ lao động pccc": "qa_pccc",
        "quần áo bảo hộ lao động an ninh thường": "qa_an_ninh_thuong",
        "quần áo bảo hộ lao động an ninh chuyên dụng": "qa_an_ninh_cd",
        "áo khoác mùa đông an ninh": "ao_khoac_dong_an_ninh",
        "quần áo chống hóa chất": "qa_chong_hoa_chat",
        "tạp dề cắt cỏ": "tap_de_cat_co",
        "áo phản quang an ninh": "ao_phan_quang_an_ninh",
        "áo phản quang nhân viên vệ sinh": "ao_phan_quang_ve_sinh",
        "áo phản quang giám sát an toàn": "ao_phan_quang_giam_sat",
        "áo mưa bộ k5 an ninh": "ao_mua_k5_an_ninh",
        "áo mưa bộ đội": "ao_mua_bo_doi",
        "dây đai an toàn": "day_dai_an_toan",
    }

    filled = 0
    for row in ws.iter_rows():
        ten_cell = row[2] if len(row) > 2 else None
        if not ten_cell or not ten_cell.value:
            continue
        ten = str(ten_cell.value).strip().lower()

        # Match product
        pid = None
        for key, mapped_id in NAME_MAP.items():
            if key in ten:
                pid = mapped_id
                break

        if not pid or pid not in products:
            continue

        prod = products[pid]
        best = prod.get("best")
        if not best or not best.get("price"):
            continue

        price = best["price"]
        # SL cell (col 7 = index 6)
        sl_cell = row[6] if len(row) > 6 else None
        sl_val = sl_cell.value if sl_cell else None
        try:
            sl = int(sl_val) if sl_val else prod.get("sl", 0)
        except:
            sl = prod.get("sl", 0)

        thanh_tien = price * sl

        # Điền đơn giá (col H = index 7)
        if len(row) > 7:
            row[7].value = price
            row[7].font = FONT
            row[7].fill = GREEN_FILL
            row[7].border = BORDER
            row[7].number_format = "#,##0"

        # Điền thành tiền (col I = index 8)
        if len(row) > 8:
            row[8].value = thanh_tien
            row[8].font = FONT
            row[8].fill = GREEN_FILL
            row[8].border = BORDER
            row[8].number_format = "#,##0"

        # Ghi chú nguồn (col J = index 9) — chỉ ghi nếu chưa có
        if len(row) > 9:
            existing_note = str(row[9].value or "")
            if "http" not in existing_note and not existing_note.startswith("Yêu cầu"):
                source_note = best.get("url", best.get("source", ""))
                hint = " [hint]" if best.get("hint_only") else ""
                scraped_date = datetime.utcnow().strftime("%d/%m/%Y")
                row[9].value = f"{source_note}{hint} (auto {scraped_date})"
                row[9].font = Font(name="Arial", size=8, italic=True, color="1F5C99")

        filled += 1
        print(f"  ✅ Filled: {ten_cell.value[:45]} → {price:,}đ × {sl} = {thanh_tien:,}đ")

    # Add metadata row at top
    ws.insert_rows(1)
    ws.merge_cells("A1:J1")
    ws["A1"].value = (
        f"Auto-filled by PPE Price Scraper | "
        f"Scraped: {price_data.get('scraped_at','')[:10]} | "
        f"Filled: {filled} products | "
        f"Source: github.com (Shopee/Tiki)"
    )
    ws["A1"].font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center")

    wb.save(excel_out)
    print(f"\nSaved to {excel_out} ({filled} cells filled)")
    return filled


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--github_url", required=True, help="GitHub raw URL to prices.json")
    parser.add_argument("--excel_in",   required=True, help="Input Excel file path")
    parser.add_argument("--excel_out",  required=True, help="Output Excel file path")
    args = parser.parse_args()

    print(f"Fetching prices from: {args.github_url}")
    price_data = fetch_prices(args.github_url)
    print(f"Got {len(price_data.get('products',{}))} products, scraped at {price_data.get('scraped_at','')}")

    filled = fill_excel(price_data, args.excel_in, args.excel_out)
    print(f"\nDone. {filled} products filled into {args.excel_out}")


if __name__ == "__main__":
    main()
